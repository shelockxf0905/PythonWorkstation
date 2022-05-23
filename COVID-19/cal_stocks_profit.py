"""
@ProjectName: 计算股票收益率
@FileName: cal_stocks_profit.py
@Author: xiao-yi.yu
@Date: 2020/09/15
"""
import openpyxl
import logging
from stocks.stock_info import stock_info


money_unit = '元'
sheet_title = ["成交日期", "证券代码", "证券名称", "股票操作", "成交价格", "成交股数", "交易金额", "手续费", "印花税", "其他杂费",
               "交易市场", "交易成本", "净利润", "净利润率"]
file_name = "D:\\yuxiaoyi\\2022-05-23\\B股历史交易明细.xlsx"


def get_stock_info_dict(code_list, stock_sheet):
    # 股票信息字典
    stock_info_dict = {}
    # 股票列表
    stock_list = []

    for row in range(2, stock_sheet.max_row + 1):
        # 交易日期
        deal_date = stock_sheet['A' + str(row)].value
        # 股票代码
        stock_code = stock_sheet['B' + str(row)].value
        # 股票名
        stock_name = stock_sheet['C' + str(row)].value
        # 股票操作
        operation = stock_sheet['D' + str(row)].value
        # 成交量
        deal_count = stock_sheet['E' + str(row)].value
        # 成交价
        deal_price = stock_sheet['F' + str(row)].value
        # 手续费
        handling_charge = stock_sheet['H' + str(row)].value
        # 印花税
        stamp_duty = stock_sheet['I' + str(row)].value
        # 其他杂费
        others = stock_sheet['J' + str(row)].value
        # 市场区分
        market_region = stock_sheet['K' + str(row)].value

        info = stock_info(deal_date, stock_code, stock_name, float(deal_price), int(deal_count),
                          operation, float(handling_charge), float(stamp_duty), float(others), market_region)
        # 股票列表作成
        stock_list.append(info)

    stocks_list = []
    info_list = []

    for code in code_list:
        for stock in stock_list:
            # 将同一股票的买卖记录抽出
            if code == str(stock.stock_code) and code not in ['700868', '700630']:
                # "龙头配股"和 "梅雁配股"需要从list中排除掉
                info_list.append(stock)

            # 这两个配股比较特殊,挑出来处理
            if stock.stock_name == "龙头配股" and '600630' == code:
                info_list.append(stock)

            if stock.stock_name == "梅雁配股" and '600868' == code:
                info_list.append(stock)

        stocks_list.append(info_list)
        info_list = []

    # 以[证券代码]为key,info为item的字典构成
    for code, stocks in zip(code_list, stocks_list):
        # 以日期为key,以同一日期的疫情数据为value
        stock_info_dict[code] = stocks

    for code in code_list:
        logging.debug('{0}的股票交易信息如下:'.format(code))
        item_list = stock_info_dict[code]
        for item in item_list:
            logging.debug('''
                交易日期={0},
                股票代码={1},
                股票名={2},
                股票操作={3},
                成交价={4}元,
                成交量={5}股,
                交易金额={6}元,
                手续费={7}元，
                印花税={8}元，
                其他杂费={9}元,
                交易市场={10}
                '''.format(item.deal_date, item.stock_code, item.stock_name, item.operation, item.deal_price,
                           item.deal_count, item.deal_price * item.deal_count, item.handling_charge, item.stamp_duty,
                           item.others, item.market_region))

    return stock_info_dict


def cal_stocks_profit():
    """
        计算股票收益
    """
    logging.debug('''----------- cal profit start -----------''')

    # 导入[A股历史成交明细.xlsx]文件
    # 导入[B股历史交易明细.xlsx]文件
    stock_wb = openpyxl.load_workbook(file_name)
    # 获取第一个sheet内容
    stock_sheet = stock_wb.worksheets[0]
    # 获取[证券代码]列内容,作为key
    code_list = []
    for row in stock_sheet.iter_rows(min_row=2, min_col=2, max_row=stock_sheet.max_row, max_col=2):
        for c in row:
            code_list.append(str(c.value))
    # 去重code列表
    code_list = sorted(set(code_list), key=code_list.index)
    i = 1
    for code in code_list:
        logging.debug('第{0}--证券代码 = {1}'.format(i, code))
        i += 1

    # 获得stock_info_dict
    stock_info_dict = get_stock_info_dict(code_list, stock_sheet)

    # 创建计算利润sheet
    cal_profit_sheet = stock_wb.create_sheet("cal_profit")

    # 设置[计算利润]sheet的第一行数据
    set_title(cal_profit_sheet)
    # 写入利润与利润率
    write_profit(cal_profit_sheet, stock_info_dict, code_list)

    # 保存工作簿
    stock_wb.save(file_name)
    logging.debug('''----------- cal profit end -------------''')


def write_profit(sheet, stock_info_dict, code_list):
    """
        计算股票收益
    """
    profit_rate = 0.0
    x = 2
    y = 1

    for code in code_list:
        item_list = stock_info_dict[code]
        for item in item_list:
            # 成交日期
            deal_date = item.deal_date
            sheet.cell(x, y, deal_date)
            y += 1
            # 证券代码
            stock_code = item.stock_code
            sheet.cell(x, y, stock_code)
            y += 1
            # 证券名称
            stock_name = item.stock_name
            sheet.cell(x, y, stock_name)
            y += 1
            # 股票操作
            operation = item.operation
            sheet.cell(x, y, operation)
            y += 1
            # 成交均价
            deal_price = item.deal_price
            sheet.cell(x, y, deal_price)
            y += 1
            # 成交数量
            deal_count = item.deal_count
            sheet.cell(x, y, deal_count)
            y += 1
            # 成交金额
            sheet.cell(x, y, deal_price * deal_count)
            y += 1
            # 手续费
            handling_charge = item.handling_charge
            sheet.cell(x, y, handling_charge)
            y += 1
            # 印花税
            stamp_duty = item.stamp_duty
            sheet.cell(x, y, stamp_duty)
            y += 1
            # 其他杂费
            others = item.others
            sheet.cell(x, y, others)
            y += 1
            # 交易市场
            sheet.cell(x, y, item.market_region)
            y += 1
            # 实际交易成本
            if operation == '买入':
                buy_cost = get_buy_cost(deal_price, deal_count, handling_charge, stamp_duty, others)
                sheet.cell(x, y, buy_cost)
                y += 1
            else:
                sell_cost = get_stock_cost(item, item_list)
                sheet.cell(x, y, sell_cost)
                y += 1
                # 交易利润
                profit = deal_price * deal_count - sell_cost
                sheet.cell(x, y, profit)
                y += 1
                # 利润率
                if sell_cost > 0:
                    profit_rate = round(profit / sell_cost * 100, 2)
                sheet.cell(x, y, "{0}%".format(profit_rate))
                y += 1
                logging.debug('''
                {0}抛售股票信息如下:
                股票代码={1},
                股票名={2},
                成交价={3},
                成交量={4}股,
                交易金额={5}
                卖出成本={6},
                卖出手续费={7}，
                卖出印花税={8}，
                卖出其他杂费={9}
                净利润={10},
                净利润率={11}%
                '''.format(deal_date, stock_code, stock_name,
                           str(deal_price) + money_unit,
                           deal_count,
                           str(deal_price * deal_count) + money_unit,
                           str(sell_cost) + money_unit,
                           str(handling_charge) + money_unit,
                           str(stamp_duty) + money_unit,
                           str(others) + money_unit,
                           str(profit) + money_unit,
                           profit_rate))
                # 计算平均利润率用
                sheet.cell(x, y, round(profit_rate/100, 2))
            y = 1
            x += 1


def get_buy_cost(deal_price, deal_count, handling_charge, stamp_duty, others):
    """
        计算购买股票的实际成本
    """
    return deal_price * deal_count + handling_charge + stamp_duty + others


def get_stock_cost(sell_item, item_list):
    cost = 0.0
    all_buy_count = 0
    all_sell_count = 0
    average_price = 0.0

    for item in item_list:
        if item.operation == '买入' and sell_item.deal_date > item.deal_date:
            # 买入
            # 卖出的日期一定大于买入的日期,中国实行n+1政策,不能当天买当天卖,至少第二天才能卖
            cost += item.deal_price * item.deal_count + item.handling_charge + item.stamp_duty + item.others
            all_buy_count += item.deal_count
        else:
            all_sell_count += item.deal_count

    if all_buy_count > 0:
        # 总买股数与总卖股数,取出最大的数来求平均价格
        average_price = round(cost / max(all_buy_count, all_sell_count), 2)

    # 当前的卖股数取得
    sell_count = sell_item.deal_count
    if sell_count == all_buy_count:
        # 买多少,卖多少
        cost += sell_item.handling_charge + sell_item.stamp_duty + sell_item.others
    elif sell_count > all_buy_count:
        # 卖多，买少
        if sell_count <= all_sell_count:
            # 卖的是总共卖的一部分,但又比买的总共要多
            # 增配股,股数增加导致当初买股单价下降,成本下降
            cost = cost / all_sell_count * all_buy_count + sell_item.handling_charge + sell_item.stamp_duty + \
                   sell_item.others
        else:
            # 卖出股数异常
            logging.error('''
            {0}抛售股票信息如下:
            股票代码={1},
            股票名={2},
            成交价={3},
            成交量={4}股
            '''.format(sell_item.deal_date, sell_item.stock_code, sell_item.stock_name,
                       str(sell_item.deal_price) + money_unit,
                       sell_item.deal_count,
                       ))
    else:
        # 卖掉部分股数
        # 卖掉的股数 * 平均价格
        cost = sell_count * average_price + sell_item.handling_charge + sell_item.stamp_duty + sell_item.others

    return cost


def set_title(sheet):
    for col in range(1, len(sheet_title) + 1):
        sheet.cell(row=1, column=col, value=sheet_title[col - 1])


def main():
    """
        主函数
    """
    cal_stocks_profit()


if __name__ == '__main__':
    main()
