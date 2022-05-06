from pyecharts.charts import Bar
from pyecharts import options as opts
import openpyxl

# 示例数据
cate = ['Apple', 'Huawei', 'Xiaomi', 'Oppo', 'Vivo', 'Meizu']
data1 = [123, 153, 89, 107, 98, 23]
data2 = [56, 77, 93, 68, 45, 67]
file_name = "D:\\yuxiaoyi\\2022-05-04\\A股历史成交明细.xlsx"


def main():
    """
        主函数
    """
    # bar = (Bar()
    #        .add_xaxis(cate)
    #        .add_yaxis('电商渠道', data1)
    #        .add_yaxis('门店', data2)
    #        .set_global_opts(title_opts=opts.TitleOpts(title="XY轴翻转-基本示例", subtitle="我是副标题"))
    #        .set_series_opts(label_opts=opts.LabelOpts(is_show=False))
    #        .reversal_axis()
    #       )
    #
    # bar.render_notebook()

    stock_wb = openpyxl.load_workbook(file_name)
    stock_sheet = stock_wb['stock_info']
    # print([[c.value for c in row] for row in stock_sheet['A2:M105']])
    #
    # print([[col.value for col in columns][1:] for columns in stock_sheet['A:K']])
    #
    # print(list(zip(*list(stock_sheet.values)[1:])))

    sh = stock_wb.worksheets[0]
    # for row in sh.rows:
    #     print([c.value for c in row])
    #
    # for col in sh.columns:
    #     print([c.value for c in col])
    #
    # print(sh.min_row, sh.min_column, sh.max_row, sh.max_column)

    # for row in sh.iter_rows(min_row=2, min_col=2, max_row=sh.max_row, max_col=2):
    #     print([c.value for c in row])

    # for row in list(sh.rows)[1:]:
    #     line = [c.value for c in row]
    #     print("[%s-%s]:%s" % (line[1], line[2], line[3:]))

    # for col in list(sh.columns)[1:]:
    #     line = [c.value for c in col][1:]
    #     print(line)

    line = [[c.value for c in row] for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row,
                                                           max_col=sh.max_column)]
    print(line)

    # line = [c.value for c in sh['B']][1:]
    # print(line)

    # total = [sum([c.value for c in row]) for row in sh.iter_rows(min_row=2, min_col=7, max_row=sh.max_row,
    # max_col=10)]
    # print(total)
    # name = [c.value for c in sh['C'][1:]]
    # print(name)
    # print(list(zip(name, total)))
    # max_data = [max([c.value for c in row]) for row in sh.iter_cols(min_row=2, min_col=7, max_row=sh.max_row,
    #                                                                 max_col=10)]
    # print(max_data)
    # name = [c.value for c in sh[1]][6:10]
    # print(name)
    # print(list(zip(name, max_data)))
    # for row in list(sh.rows)[1:]:
    #     line = [v.value for v in row]
    #     print(line[2], sum(line[6:9]))

    # for col in list(sh.columns):
    #     column_data = [v.value for v in col]
    #
    # print(column_data)

    # 工作表['A1'] =
    # 工作表.cell(2, 3, 'Gandum is comming.')

    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = "十九乘法口诀表"
    # for x in range(1, 20):
    #     for y in range(1, x + 1):
    #         print(y, '*', x, '=', y*x, end='\t')
    #         ws.cell(x, y, '%d × %d = %d' % (y, x, y*x))
    #     print('')
    # wb.save('十九乘法表.xlsx')

    # for r in [['%d × %d = %d' % (y, x, x * y) for y in range(1, x + 1)] for x in range(1, 20)]:
    #     ws.append(r)
    # wb.save('十九乘法表.xlsx')


if __name__ == '__main__':
    main()
